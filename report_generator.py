"""
报告生成模块 - 将筛选结果输出为 Excel 报告
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import MatchResult


# 样式定义
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_RECOMMEND_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_MAYBE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_REJECT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel_report(results: list[MatchResult], output_dir: str = None) -> str:
    """
    生成 Excel 筛选报告
    
    Args:
        results: 匹配结果列表（已按分数排序）
        output_dir: 输出目录，默认使用 config 中的配置
        
    Returns:
        生成的 Excel 文件路径
    """
    if output_dir is None:
        output_dir = config.OUTPUT_DIR

    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    
    # Sheet 1: 总览
    _create_overview_sheet(wb, results)
    
    # Sheet 2: 详细评分
    _create_detail_sheet(wb, results)

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"简历筛选报告_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath


def _create_overview_sheet(wb: Workbook, results: list[MatchResult]):
    """创建总览 Sheet"""
    ws = wb.active
    ws.title = "筛选总览"

    # 标题行
    headers = ["排名", "姓名", "总分", "推荐等级", "学历", "工作年限", "优势", "不足", "总评"]
    _write_header_row(ws, headers)

    # 数据行
    for i, result in enumerate(results):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=result.candidate_name)
        ws.cell(row=row, column=3, value=result.total_score)
        ws.cell(row=row, column=4, value=result.recommendation)
        ws.cell(row=row, column=5, value=result.resume.education or "未知")
        ws.cell(row=row, column=6, value=result.resume.work_years or "未知")
        ws.cell(row=row, column=7, value=result.strengths)
        ws.cell(row=row, column=8, value=result.weaknesses)
        ws.cell(row=row, column=9, value=result.overall_comment)

        # 根据推荐等级设置行背景色
        if result.total_score >= config.RECOMMEND_THRESHOLD:
            fill = _RECOMMEND_FILL
        elif result.total_score >= config.MAYBE_THRESHOLD:
            fill = _MAYBE_FILL
        else:
            fill = _REJECT_FILL

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 调整列宽
    col_widths = [6, 12, 8, 14, 8, 10, 35, 35, 40]
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _create_detail_sheet(wb: Workbook, results: list[MatchResult]):
    """创建详细评分 Sheet"""
    ws = wb.create_sheet("详细评分")

    # 动态表头：基础列 + 各评分维度
    dim_names = [d["name"] for d in config.SCORING_DIMENSIONS]
    headers = ["排名", "姓名", "总分", "推荐等级"] + dim_names + ["文件名"]
    _write_header_row(ws, headers)

    # 数据行
    for i, result in enumerate(results):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=result.candidate_name)
        ws.cell(row=row, column=3, value=result.total_score)
        ws.cell(row=row, column=4, value=result.recommendation)

        # 各维度分数
        dim_score_map = {d.name: d for d in result.dimensions}
        for j, dim_name in enumerate(dim_names):
            dim = dim_score_map.get(dim_name)
            score_text = f"{dim.score}" if dim else "N/A"
            ws.cell(row=row, column=5 + j, value=score_text)

        ws.cell(row=row, column=5 + len(dim_names), value=result.resume.file_name)

        # 样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")

    # 调整列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    for i in range(len(dim_names)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 12
    ws.column_dimensions[get_column_letter(5 + len(dim_names))].width = 30


def _write_header_row(ws, headers: list[str]):
    """写入格式化的表头行"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def generate_console_report(results: list[MatchResult]):
    """在终端打印简要报告"""
    from rich.console import Console
    from rich.table import Table

    console = Console()

    # 统计
    recommend = sum(1 for r in results if r.total_score >= config.RECOMMEND_THRESHOLD)
    maybe = sum(1 for r in results if config.MAYBE_THRESHOLD <= r.total_score < config.RECOMMEND_THRESHOLD)
    reject = sum(1 for r in results if r.total_score < config.MAYBE_THRESHOLD)

    console.print(f"\n📊 [bold]筛选结果统计[/bold]")
    console.print(f"   总计: {len(results)} 人 | ⭐推荐: {recommend} 人 | 🔶待定: {maybe} 人 | ❌不推荐: {reject} 人\n")

    # 表格
    table = Table(title="候选人排名", show_lines=True)
    table.add_column("排名", style="bold", width=4, justify="center")
    table.add_column("姓名", width=10)
    table.add_column("总分", width=6, justify="center")
    table.add_column("等级", width=12)
    table.add_column("总评", width=50)

    for i, result in enumerate(results):
        # 根据分数设置颜色
        if result.total_score >= config.RECOMMEND_THRESHOLD:
            style = "green"
        elif result.total_score >= config.MAYBE_THRESHOLD:
            style = "yellow"
        else:
            style = "red"

        table.add_row(
            str(i + 1),
            result.candidate_name,
            f"[{style}]{result.total_score}[/{style}]",
            result.recommendation,
            result.overall_comment[:80] if result.overall_comment else (result.error or ""),
        )

    console.print(table)
